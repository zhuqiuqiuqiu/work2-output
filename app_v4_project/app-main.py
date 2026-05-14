# -*- coding: utf-8 -*-
import pandas as pd
import os
import glob
import tkinter as tk
from tkinter import filedialog, messagebox


class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("配料清单批量生成工具")
        self.root.geometry("500x400")
        # 要零件material part—name 和vender数据
        tk.Label(root, text="零件-零件名-供应商数据").pack(pady=5)
        self.master_entry = tk.Entry(root, width=50)
        self.master_entry.pack()
        tk.Button(root, text="选择Excel文件", command=self.select_master_file).pack(pady=5)
        # 输入文件夹
        tk.Label(root, text="输入文件夹:").pack(pady=5)
        self.input_entry = tk.Entry(root, width=50)
        self.input_entry.pack()
        tk.Button(root, text="选择文件夹", command=self.select_input_folder).pack(pady=5)
        # 输出文件夹
        tk.Label(root, text="输出文件夹:").pack(pady=5)
        self.output_entry = tk.Entry(root, width=50)
        self.output_entry.pack()
        tk.Button(root, text="选择文件夹", command=self.select_output_folder).pack(pady=5)
        # 开始处理按钮
        tk.Button(root, text="生成料车配料清单", width=20, command=self.process_files, bg="#4CAF50", fg="red").pack(
            pady=20)

    def select_input_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.input_entry.delete(0, tk.END)
            self.input_entry.insert(0, folder)

    def select_output_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, folder)

    def select_master_file(self):
        file_path = filedialog.askopenfilename(
            title="选择主数据表",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            self.master_entry.delete(0, tk.END)
            self.master_entry.insert(0, file_path)

    def process_files(self):
        input_folder = self.input_entry.get().strip()
        output_folder = self.output_entry.get().strip()
        master_path = self.master_entry.get().strip()
        if not master_path or not os.path.exists(master_path):
            messagebox.showerror("错误", "请选择有效的主数据表（包含 Material / part-name / 供应商）")
            return

        df_master = pd.read_excel(master_path)
        df_master.columns = df_master.columns.str.strip()

        # 校验必要列
        required_cols = {"Material", "part_name", "供应商"}
        if not required_cols.issubset(set(df_master.columns)):
            messagebox.showerror(
                "错误",
                f"主数据表缺少必要列：{required_cols - set(df_master.columns)}"
            )
            return

        # 关键修复：去重，确保每个 Material 只有一条记录
        df_master = df_master[["Material", "part_name", "供应商"]].drop_duplicates(subset=["Material"], keep="first")

        if not input_folder or not os.path.exists(input_folder):
            messagebox.showerror("错误", "请输入有效的输入文件夹！")
            return
        if not output_folder:
            output_folder = os.path.join(input_folder, "output_excel")
        os.makedirs(output_folder, exist_ok=True)

        files = glob.glob(os.path.join(input_folder, "*.xlsx"))
        if not files:
            messagebox.showwarning("提示", "输入文件夹没有找到 Excel 文件！")
            return

        for file_path in files:
            try:
                df = pd.read_excel(file_path)
                df.columns = df.columns.str.strip()

                # 汇总数量
                df_qty = df.groupby(["Material", "Serial Number"], as_index=False)["QTY"].sum()
                # 唯一 Bin
                df_bin = df.groupby("Material")["Storage Bin"].agg(
                    lambda x: ",".join(sorted(set(map(str, x))))
                ).reset_index()
                # 数量矩阵
                qty_matrix = pd.pivot_table(df_qty, index="Material", columns="Serial Number", values="QTY",
                                            fill_value=0)
                result = qty_matrix.map(lambda x: f"{int(x)}" if x > 0 else "×")
                # 合并 Bin
                result = result.reset_index().merge(df_bin, on="Material", how="left")
                # 合并 part-name 和 供应商
                result = result.merge(df_master, on="Material", how="left")

                # ========== 功能 1: 根据 Storage Bin 按原始字符串排序 ==========
                result = result.sort_values(by='Storage Bin', ascending=True, na_position='last')
                # 重置索引
                result = result.reset_index(drop=True)

                # ========== 功能 2: 添加序号列 ==========
                result.insert(0, '序号', range(1, len(result) + 1))

                # ========== 功能 4: 清理 part_name 列（去掉"，"和"；"后面的文本） ==========
                def clean_part_name(name):
                    if pd.isna(name):
                        return name
                    # 先按"，"分割取第一部分，再按"；"分割取第一部分
                    name_str = str(name)
                    name_str = name_str.split('，')[0]  # 中文逗号
                    name_str = name_str.split(',')[0]  # 英文逗号
                    name_str = name_str.split('；')[0]  # 中文分号
                    name_str = name_str.split(';')[0]  # 英文分号
                    return name_str.strip()

                if 'part_name' in result.columns:
                    result['part_name'] = result['part_name'].apply(clean_part_name)

                # 调整列顺序
                sn_cols = [c for c in result.columns
                           if c not in ["序号", "Material", "part_name", "供应商", "Storage Bin"]]

                cols = ["序号", "Material", "part_name", "供应商"] + ["Storage Bin"] + sn_cols
                # 确保所有列都存在
                cols = [c for c in cols if c in result.columns]
                result = result[cols]

                # ========== 功能 3: 添加标题行（在列名之前插入） ==========
                # 获取文件名（不含扩展名）
                file_name = os.path.basename(file_path)
                file_name_without_ext = os.path.splitext(file_name)[0]
                # 提取第一个"-"前面的文本
                title_text = file_name_without_ext.split('-')[
                    0] if '-' in file_name_without_ext else file_name_without_ext

                # 保存列名
                original_columns = result.columns.tolist()

                # 将列名转换为数据行（第二行）
                # 创建标题行数据（第一行）
                title_row_data = [title_text] + [''] * (len(original_columns) - 1)

                # 创建包含标题行和列名行的DataFrame
                header_df = pd.DataFrame([title_row_data, original_columns], columns=original_columns)

                # 合并标题行、列名行和数据
                result = pd.concat([header_df, result], ignore_index=True)

                output_path = os.path.join(output_folder, os.path.basename(file_path))
                # 不写入列名（因为已经在数据中了），不写入索引
                result.to_excel(output_path, index=False, header=False)

                # 增加文字样式
                from openpyxl import load_workbook
                from openpyxl.styles import Font, Alignment
                # 重新打开 Excel 进行样式处理
                wb = load_workbook(output_path)
                ws = wb.active
                # 设置第一行样式
                title_cell = ws.cell(row=1, column=1)
                title_cell.font = Font(size=16, bold=True)  # 字号16 + 加粗
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                # 第二行表头加粗
                from openpyxl.styles import PatternFill, Font

                header_fill = PatternFill("solid", fgColor="E7E6E6")
                header_font = Font(bold=True)

                for cell in ws[2]:
                    cell.fill = header_fill
                    cell.font = header_font
                # 合并第一行（横向铺满整行）
                ws.merge_cells(
                    start_row=1, start_column=1,
                    end_row=1, end_column=ws.max_column
                )
                from openpyxl.styles import Border, Side, Alignment
                # 细边框样式
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                # 全表居中 + 加边框
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                        min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border

                #自带列宽
                from openpyxl.utils import get_column_letter

                for col_idx in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    max_length = 0

                    for row in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_idx)

                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))

                    ws.column_dimensions[col_letter].width = max_length + 2
###########################################################################
                # ========== A4 打印适配 ==========
                from openpyxl.worksheet.worksheet import Worksheet

                # 页面设置为 A4
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # 纵向
                # 如果你列很多可改为横向：
                # ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

                # 页面居中
                ws.page_setup.horizontalCentered = True
                ws.page_setup.verticalCentered = True

                # 页边距（适合打印）
                ws.page_margins.left = 0.5
                ws.page_margins.right = 0.5
                ws.page_margins.top = 0.75
                ws.page_margins.bottom = 0.75

                # 自动缩放到一页宽（核心）
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = False

                # 每页重复打印：标题 + 表头
                ws.print_title_rows = '1:2'

                # 设置打印区域
                ws.print_area = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"

                # 自动行高（适合打印）
                def auto_adjust_row_height(ws, min_height=20, max_height=60):
                    for row in ws.iter_rows():
                        max_len = 0
                        row_idx = row[0].row

                        for cell in row:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))

                        height = min(max(min_height, max_len * 1.5), max_height)
                        ws.row_dimensions[row_idx].height = height

                auto_adjust_row_height(ws)
################################################################
                wb.save(output_path)




            except Exception as e:
                messagebox.showerror("处理错误", f"文件 {os.path.basename(file_path)} 处理失败:\n{e}")
                continue

        messagebox.showinfo("完成", f"所有文件已处理完成！\n输出文件夹:\n{output_folder}")


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()